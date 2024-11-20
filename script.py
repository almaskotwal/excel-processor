import openpyxl
import os

def process_excel(input_file, template_file, output_dir):
  """Processes an Excel file, populates templates, and saves new files.

  Args:
      input_file: Path to the input Excel file.
      template_file: Path to the template Excel file.
      output_dir: Path to the output directory.
  """

  workbook = openpyxl.load_workbook(input_file)
  worksheet = workbook.active

  template_workbook = openpyxl.load_workbook(template_file)
  template_worksheet = template_workbook.active

  for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
    if row:
      # Extract values from the specified columns
      Trip_ID, Driver_Name = row[:2]
      template_worksheet['B11'] = Trip_ID
      template_worksheet['D4'] = Driver_Name
      

      # Save the modified template as a new file using the "bill_to" value
      output_file = os.path.join(output_dir, f"{Driver_Name}.xlsx")
      template_workbook.save(output_file)

# Replace with your actual file paths
input_file = "input/trips.xlsx"
template_file = "template/payroll_template.xlsx"
output_dir = "output_files"

process_excel(input_file, template_file, output_dir)
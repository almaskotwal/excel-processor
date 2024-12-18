import pytz
import streamlit as st
import openpyxl
import os
from datetime import datetime, timedelta
import zipfile

def process_excel(input_file, template_file, output_dir):
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active

    target_columns = {"Trip ID": None, "Driver Name": None, "Facility Sequence": None, "Estimated Cost": None}

    for header_cell in worksheet.iter_rows(min_row=2, values_only=True):
        if header_cell:
            for col_idx, value in enumerate(header_cell):
                if value is not None:
                    value_str = str(value)
                    if value_str.strip() in target_columns:
                        target_columns[value_str.strip()] = col_idx + 1

    if not all(value for value in target_columns.values()):
        raise ValueError(f"Missing target columns in input file: {', '.join(missing for missing in target_columns if not target_columns[missing])}")

    processed_drivers = {}

    start_row = 16
    facility_row = 18
    estimated_cost_row = 19
    cell_offset = 5

    for row_num, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), 3):
        if row:
            Trip_ID = row[target_columns["Trip ID"] - 1]
            Driver_Name = row[target_columns["Driver Name"] - 1]
            Facility_Sequence = row[target_columns["Facility Sequence"] - 1]
            Estimated_Cost = row[target_columns["Estimated Cost"] - 1]

            output_file = os.path.join(output_dir, f"{Driver_Name}.xlsx")

            if Driver_Name not in processed_drivers:
                processed_drivers[Driver_Name] = {'trip_row': start_row, 'facility_row': facility_row, 'estimated_cost_row': estimated_cost_row}
                template_workbook = openpyxl.load_workbook(template_file)
                template_worksheet = template_workbook.active

               # Calculate start and end dates in Mountain Time
                mountain_tz = pytz.timezone('MST')
                today = datetime.now(mountain_tz)
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)

                            # Adjust to the nearest Saturday
                if today.weekday() == 6:  # If today is Saturday
                    start_date = today
                else:
                    start_date = today - timedelta(days=today.weekday() + 1)

                end_date = start_date + timedelta(days=6)

                # Populate date fields in the template
                template_worksheet['D3'] = today.strftime('%m/%d/%Y')
                template_worksheet['D6'] = start_date.strftime('%m/%d/%Y')
                template_worksheet['D7'] = end_date.strftime('%m/%d/%Y')

                template_worksheet['B11'] = Trip_ID
                template_worksheet['D4'] = Driver_Name
                template_worksheet['B13'] = Facility_Sequence
                template_worksheet['B14'] = Estimated_Cost
                template_workbook.save(output_file)
            else:
                existing_workbook = openpyxl.load_workbook(output_file)
                driver_data = processed_drivers[Driver_Name]
                current_trip_row = driver_data['trip_row']
                current_facility_row = driver_data['facility_row']
                current_estimated_cost_row = driver_data['estimated_cost_row']

                existing_worksheet = existing_workbook.active
                existing_worksheet[f"B{current_trip_row}"] = Trip_ID
                existing_worksheet[f"B{current_facility_row}"] = Facility_Sequence
                existing_worksheet[f"B{current_estimated_cost_row}"] = Estimated_Cost

                driver_data['trip_row'] += cell_offset
                driver_data['facility_row'] += cell_offset
                driver_data['estimated_cost_row'] += cell_offset

                existing_workbook.save(output_file)

    return processed_drivers

def main():
    st.title("Excel Processor")

    uploaded_input_file = st.file_uploader("Upload Input File", type=["xlsx"])
    uploaded_template_file = st.file_uploader("Upload Template File", type=["xlsx"])

    output_dir = st.text_input("Output Directory (Leave blank for default)", value="output_files")

    if st.button("Process"):
        if uploaded_input_file is not None and uploaded_template_file is not None:
            with open("input.xlsx", "wb") as f:
                f.write(uploaded_input_file.read())
            with open("template.xlsx", "wb") as f:
                f.write(uploaded_template_file.read())

            processed_drivers = process_excel("input.xlsx", "template.xlsx", output_dir)

            # Create a ZIP file containing all output files
            zip_filename = "output_files.zip"
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for driver_name in processed_drivers:
                    output_file = os.path.join(output_dir, f"{driver_name}.xlsx")
                    zipf.write(output_file)

            with open(zip_filename, "rb") as f:
                st.download_button("Download All Files", f, file_name=zip_filename)

            st.success("Processing complete! Download the ZIP file above.")
        else:
            st.warning("Please upload both input and template files.")

if __name__ == "__main__":
    main()
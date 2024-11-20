import streamlit as st
import openpyxl
import io

def process_excel(uploaded_input_file, uploaded_template_file, output_dir):
    """Processes an Excel file, populates templates, and creates new files.

    Args:
        uploaded_input_file: Uploaded input file object (Streamlit.UploadedFile).
        uploaded_template_file: Path to the template Excel file.
        output_dir: Path to the output directory (ignored in this version).
    """

    try:
        # Read uploaded files into memory
        input_data = uploaded_input_file.read()
        template_data = uploaded_template_file.read()

        # Load workbooks from memory using BytesIO
        input_workbook = openpyxl.load_workbook(io.BytesIO(input_data))
        worksheet = input_workbook.active

        # Define the column names we want to extract, matching the actual header names
        target_columns = {"Trip ID": None, "Driver Name": None, "Facility Sequence": None, "Estimated Cost": None}

        # Find the column index for each target column name, considering headers are on row 2
        for header_cell in worksheet.iter_rows(min_row=2, values_only=True):
            if header_cell:
                for col_idx, value in enumerate(header_cell):
                    if value is not None:
                        value_str = str(value)
                        if value_str.strip() in target_columns:
                            target_columns[value_str.strip()] = col_idx + 1

        # Create output workbook in memory
        output_workbook = openpyxl.Workbook()
        output_worksheet = output_workbook.active

        # Populate output worksheet with processed data
        # Check if all target columns were found
        if not all(value for value in target_columns.values()):
            raise ValueError(f"Missing target columns in input file: {', '.join(missing for missing in target_columns if not target_columns[missing])}")

        processed_drivers = {}  # To keep track of processed drivers, next row for each

        start_row = 16  # Starting row for Trip IDs
        facility_row = 18  # Starting row for Facility Sequence
        estimated_cost_row = 19  # Starting row for Estimated Cost
        cell_offset = 5  # Increment for each new Trip ID, Facility Sequence, and Estimated Cost

        for row_num, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), 3):
            if row:
                Trip_ID = row[target_columns["Trip ID"] - 1]
                Driver_Name = row[target_columns["Driver Name"] - 1]
                Facility_Sequence = row[target_columns["Facility Sequence"] - 1]
                Estimated_Cost = row[target_columns["Estimated Cost"] - 1]

                output_file = os.path.join(output_dir, f"{Driver_Name}.xlsx")

                if Driver_Name not in processed_drivers:
                    # First occurrence of the driver
                    processed_drivers[Driver_Name] = {'trip_row': start_row, 'facility_row': facility_row, 'estimated_cost_row': estimated_cost_row}
                    template_workbook = openpyxl.load_workbook(template_file)
                    template_worksheet = template_workbook.active

                    template_worksheet['D4'] = Driver_Name
                    template_worksheet['B11'] = Trip_ID
                    template_worksheet['B13'] = Facility_Sequence
                    template_worksheet['B14'] = Estimated_Cost
                    template_workbook.save(output_file)
                else:
                    # Subsequent occurrences of the driver
                    existing_workbook = openpyxl.load_workbook(output_file)
                    driver_data = processed_drivers[Driver_Name]
                    current_trip_row = driver_data['trip_row']
                    current_facility_row = driver_data['facility_row']
                    current_estimated_cost_row = driver_data['estimated_cost_row']

                    existing_worksheet = existing_workbook.active
                    existing_worksheet[f"B{current_trip_row}"] = Trip_ID
                    existing_worksheet[f"B{current_facility_row}"] = Facility_Sequence
                    existing_worksheet[f"B{current_estimated_cost_row}"] = Estimated_Cost

                    # Update rows for the next iteration with offset
                    driver_data['trip_row'] += cell_offset
                    driver_data['facility_row'] += cell_offset
                    driver_data['estimated_cost_row'] += cell_offset

                    existing_workbook.save(output_file)


        # Create a temporary file object in memory for download
        output_data = io.BytesIO()
        output_workbook.save(output_data)
        output_data.seek(0)

        # Generate a descriptive filename based on processed data (e.g., driver name)
        output_filename = f"{processed_drivers['Driver Name']}.xlsx"  # Replace with your desired filename

        # Download the processed file using Streamlit download button
        st.download_button(label="Download Processed File", data=output_data, file_name=output_filename)

    except FileNotFoundError:
        st.error(f"Failed to open input file.")

def main():
    st.title("Excel Processor")

    # File upload
    uploaded_input_file = st.file_uploader("Upload Input File", type=["xlsx"])
    uploaded_template_file = st.file_uploader("Upload Template File", type=["xlsx"])

    if st.button("Process"):
        if uploaded_input_file is not None and uploaded_template_file is not None:
            process_excel(uploaded_input_file, uploaded_template_file)  # Removed output_dir
        else:
            st.warning("Please upload both input and template files.")

if __name__ == "__main__":
    main()
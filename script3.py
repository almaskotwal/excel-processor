import openpyxl
import os

def process_excel(input_file, template_file, output_dir):
    """Processes an Excel file, populates templates, and saves new files.

    Args:
        input_file: Path to the input Excel file.
        template_file: Path to the template Excel file.
        output_dir: Path to the output directory.
    """

    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active

    # Define the column names we want to extract, matching the actual header names
    target_columns = {"Trip ID": None, "Driver Name": None}

    # Find the column index for each target column name, considering headers are on row 2
    for header_cell in worksheet.iter_rows(min_row=2, values_only=True):
        if header_cell:
            for col_idx, value in enumerate(header_cell):
                if value is not None:
                    value_str = str(value)
                    if value_str.strip() in target_columns:
                        target_columns[value_str.strip()] = col_idx + 1

    # Check if all target columns were found
    if not all(value for value in target_columns.values()):
        raise ValueError(f"Missing target columns in input file: {', '.join(missing for missing in target_columns if not target_columns[missing])}")

    processed_drivers = {}  # To keep track of processed drivers and next row for each

    for row_num, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), 3):
        if row:
            Trip_ID = row[target_columns["Trip ID"] - 1]
            Driver_Name = row[target_columns["Driver Name"] - 1]

            output_file = os.path.join(output_dir, f"{Driver_Name}.xlsx")

            if Driver_Name not in processed_drivers:
                # First occurrence of the driver
                processed_drivers[Driver_Name] = 16  # Initialize next row to B16
                template_workbook = openpyxl.load_workbook(template_file)
                template_worksheet = template_workbook.active
                template_worksheet['B11'] = Trip_ID
                template_worksheet['D4'] = Driver_Name
                template_workbook.save(output_file)
            else:
                # Subsequent occurrences of the driver
                existing_workbook = openpyxl.load_workbook(output_file)
                existing_worksheet = existing_workbook.active
                next_row = processed_drivers[Driver_Name]

                # Check if we've reached the limit of 5 rows
                if next_row <= 36:
                    existing_worksheet[f"B{next_row}"] = Trip_ID
                    processed_drivers[Driver_Name] += 5
                else:
                    print(f"Reached maximum capacity for {Driver_Name}. Skipping additional Trip IDs.")

                existing_workbook.save(output_file)

# Replace with your actual file paths
input_file = "input/trips.xlsx"
template_file = "template/payroll_template.xlsx"
output_dir = "output_files"

process_excel(input_file, template_file, output_dir)